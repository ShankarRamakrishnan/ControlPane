import json
import os
from datetime import datetime
from pathlib import Path
from langchain_core.tools import tool
from gateway.core.tool_registry import register


@register
@tool
def export_to_excel(properties_json: str, filename: str = "") -> str:
    """Export a list of properties to an Excel (.xlsx) file.

    Args:
        properties_json: JSON string — either a list of property dicts, or an
                         object with a 'properties' key containing that list.
                         Each property should have keys like address, price,
                         beds, baths, sqft, price_per_sqft, days_on_market,
                         zestimate, url, etc.
        filename: Optional output filename (without path). Defaults to
                  'zillow_results_<timestamp>.xlsx'. The file is saved to
                  the OUTPUT_DIR env var (default: current working directory).

    Returns:
        Absolute path of the saved Excel file, or an error message.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "Error: openpyxl is not installed. Run: pip install openpyxl"

    # Parse input
    try:
        data = json.loads(properties_json)
    except json.JSONDecodeError as e:
        return f"Error: invalid JSON input — {e}"

    if isinstance(data, dict):
        properties = data.get("properties", [])
    elif isinstance(data, list):
        properties = data
    else:
        return "Error: expected a JSON list or object with a 'properties' key."

    if not properties:
        return "No properties to export."

    # Column definitions: (header label, data key)
    columns = [
        ("Address", "address"),
        ("Price ($)", "price"),
        ("Beds", "beds"),
        ("Baths", "baths"),
        ("Sqft", "sqft"),
        ("$/Sqft", "price_per_sqft"),
        ("Lot Size (sqft)", "lot_size_sqft"),
        ("Home Type", "home_type"),
        ("Year Built", "year_built"),
        ("Days on Market", "days_on_market"),
        ("Zestimate ($)", "zestimate"),
        ("Rent Zestimate ($)", "rent_zestimate"),
        ("HOA Fee ($)", "hoa_fee"),
        ("Status", "listing_status"),
        ("URL", "url"),
        ("Passes Buybox", "passes_buybox"),
        ("Buybox Notes", "buybox_notes"),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Properties"

    # Header row styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (label, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 30

    # Alternate row fills
    even_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    pass_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    fail_fill = PatternFill(fill_type="solid", fgColor="FFCCCC")

    for row_idx, prop in enumerate(properties, start=2):
        passes = prop.get("passes_buybox")
        row_fill = pass_fill if passes is True else (fail_fill if passes is False else (even_fill if row_idx % 2 == 0 else None))

        for col_idx, (_, key) in enumerate(columns, start=1):
            value = prop.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_fill:
                cell.fill = row_fill
            if key == "url" and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")
            if key in ("price", "zestimate", "rent_zestimate", "hoa_fee") and isinstance(value, (int, float)):
                cell.number_format = '#,##0'
            if key == "price_per_sqft" and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'

    # Auto-fit columns
    for col_idx, (label, _) in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_width = len(label)
        for row_idx in range(2, len(properties) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_width = max(max_width, min(len(str(val)), 50))
        ws.column_dimensions[col_letter].width = max_width + 2

    # Freeze header row
    ws.freeze_panes = "A2"

    # Summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    passed = [p for p in properties if p.get("passes_buybox") is True]
    ws_summary["A1"] = "Buybox Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = "Total Listings Searched"
    ws_summary["B3"] = len(properties)
    ws_summary["A4"] = "Passed Buybox"
    ws_summary["B4"] = len(passed)
    ws_summary["A5"] = "Pass Rate"
    ws_summary["B5"] = f"{len(passed)/len(properties)*100:.1f}%" if properties else "N/A"
    if passed:
        prices = [p["price"] for p in passed if isinstance(p.get("price"), (int, float))]
        ws_summary["A7"] = "Avg Price (Passed)"
        ws_summary["B7"] = round(sum(prices) / len(prices)) if prices else "N/A"
        ws_summary["A8"] = "Min Price (Passed)"
        ws_summary["B8"] = min(prices) if prices else "N/A"
        ws_summary["A9"] = "Max Price (Passed)"
        ws_summary["B9"] = max(prices) if prices else "N/A"
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 20

    # Determine output path
    output_dir = Path(os.getenv("OUTPUT_DIR", ".")).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"zillow_results_{timestamp}.xlsx"
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    output_path = output_dir / filename

    wb.save(str(output_path))
    return str(output_path)
